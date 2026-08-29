"""Tabular profiling for coverage and staleness — CSV, TSV, and XLSX."""

from __future__ import annotations

import csv
import hashlib
from pathlib import Path

TABULAR_SUFFIXES = {".csv", ".tsv", ".xlsx"}


def profile_file(path: Path) -> dict | None:
    """Return a facts-shaped dict compatible with `staleness.extract_measures`, or None."""
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _profile_xlsx(path)
    if suffix in {".csv", ".tsv"}:
        return _profile_csv(path, delimiter="\t" if suffix == ".tsv" else ",")
    return None


def file_sha256(path: Path) -> str:
    """Hash file contents, or empty string on failure."""
    try:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(65536), b""):
                digest.update(chunk)
        return digest.hexdigest()
    except OSError:
        return ""


def _profile_csv(path: Path, *, delimiter: str) -> dict | None:
    try:
        with path.open(newline="", encoding="utf-8") as handle:
            reader = csv.DictReader(handle, delimiter=delimiter)
            fieldnames = reader.fieldnames or []
            rows = list(reader)
    except (OSError, UnicodeDecodeError, csv.Error):
        return None

    if not fieldnames:
        return {"rows": 0, "numeric": []}

    return _facts_from_rows(fieldnames, rows)


def _profile_xlsx(path: Path) -> dict | None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            workbook.close()
            return {"rows": 0, "numeric": []}
        fieldnames = [str(cell) if cell is not None else "" for cell in header]
        rows = [
            {fieldnames[i]: ("" if value is None else value) for i, value in enumerate(row)}
            for row in rows_iter
        ]
        workbook.close()
    except (OSError, ValueError):
        return None

    return _facts_from_rows(fieldnames, rows)


def _facts_from_rows(fieldnames: list[str], rows: list[dict]) -> dict:
    numeric: list[dict[str, object]] = []
    row_count = len(rows)
    for column in fieldnames:
        values: list[float] = []
        for row in rows:
            raw = str(row.get(column) or "").strip()
            if not raw:
                continue
            cleaned = raw.replace(",", "").replace("$", "").replace("£", "").replace("€", "")
            if cleaned.endswith("%"):
                cleaned = cleaned[:-1]
            try:
                values.append(float(cleaned))
            except ValueError:
                values = []
                break
        if values and len(values) >= max(1, int(row_count * 0.8)):
            numeric.append({"name": column, "total": sum(values)})

    return {"rows": row_count, "numeric": numeric}
